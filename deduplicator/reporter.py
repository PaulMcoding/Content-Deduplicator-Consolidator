from pathlib import Path

import pandas as pd
from openpyxl.styles import Font, PatternFill

from .models import DuplicateGroup


def build_report_dataframe(groups: list[DuplicateGroup]) -> pd.DataFrame:
    """Convert DuplicateGroups into a flat DataFrame for reporting."""
    rows = []
    for group in groups:
        for f in group.files:
            is_recommended = (
                group.recommended is not None
                and f.path == group.recommended.path
            )
            rows.append({
                "Relative Path": group.relative_path,
                "Root Folder": f.root_folder,
                "Full Path": str(f.path),
                "Size (MB)": round(f.size / (1024 * 1024), 2),
                "Last Modified": f.modified.strftime("%Y-%m-%d %H:%M:%S"),
                "SHA-256": f.sha256[:16] + "...",
                "SHA-256 Full": f.sha256,
                "Category": group.category,
                "Recommended": "YES" if is_recommended else "",
            })

    df = pd.DataFrame(rows)
    if not df.empty:
        category_order = ["conflicting_versions", "exact_duplicate", "unique"]
        df["_sort"] = df["Category"].map(
            {c: i for i, c in enumerate(category_order)}
        )
        df = df.sort_values(["_sort", "Relative Path", "Root Folder"])
        df = df.drop(columns=["_sort"])
        df = df.reset_index(drop=True)
    return df


def export_to_csv(df: pd.DataFrame, output_path: Path) -> None:
    """Export report DataFrame to CSV."""
    export_df = df.drop(columns=["SHA-256 Full"], errors="ignore")
    export_df.to_csv(output_path, index=False)


def export_to_excel(df: pd.DataFrame, output_path: Path) -> None:
    """Export report to multi-sheet Excel with formatting."""
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        # Full report
        export_df = df.drop(columns=["SHA-256 Full"], errors="ignore")
        export_df.to_excel(writer, sheet_name="All Files", index=False)

        # Per-category sheets
        for category, label in [
            ("conflicting_versions", "Conflicts - REVIEW"),
            ("exact_duplicate", "Exact Duplicates"),
            ("unique", "Unique Files"),
        ]:
            cat_df = export_df[export_df["Category"] == category]
            if not cat_df.empty:
                cat_df.to_excel(writer, sheet_name=label, index=False)

        # Apply formatting
        wb = writer.book
        fills = {
            "Conflicts - REVIEW": PatternFill(
                start_color="FFE0B2", end_color="FFE0B2", fill_type="solid"
            ),
            "Exact Duplicates": PatternFill(
                start_color="C8E6C9", end_color="C8E6C9", fill_type="solid"
            ),
            "Unique Files": PatternFill(
                start_color="BBDEFB", end_color="BBDEFB", fill_type="solid"
            ),
        }

        for sheet_name, fill in fills.items():
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                # Bold header row
                for cell in ws[1]:
                    cell.font = Font(bold=True)
                # Color data rows
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                    for cell in row:
                        cell.fill = fill
                # Auto-width columns
                for col in ws.columns:
                    max_len = max(
                        len(str(cell.value or "")) for cell in col
                    )
                    ws.column_dimensions[col[0].column_letter].width = min(
                        max_len + 2, 60
                    )

        # Format "All Files" sheet header
        if "All Files" in wb.sheetnames:
            ws = wb["All Files"]
            for cell in ws[1]:
                cell.font = Font(bold=True)
            for col in ws.columns:
                max_len = max(
                    len(str(cell.value or "")) for cell in col
                )
                ws.column_dimensions[col[0].column_letter].width = min(
                    max_len + 2, 60
                )
